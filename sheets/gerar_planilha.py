# -*- coding: utf-8 -*-
"""Gera Controle-Atos-Cartorio.xlsx para importar no Google Sheets.

Sem Apps Script: tudo que a planilha faz sozinha vem de formulas nativas.
Formulas sao gravadas com nomes de funcao em ingles e virgula como separador
(e o que o formato .xlsx exige); o Sheets traduz na exibicao para pt-BR.
"""
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule

LINHAS = 301  # dados nas linhas 2 a 301

# ---------------------------------------------------------------- dados base
TIPOS = [
    ('Procuração Ad Judicia', 'Procurações'),
    ('Procuração Ad Negotia', 'Procurações'),
    ('Procuração para venda de imóvel', 'Procurações'),
    ('Procuração bancária/financeira', 'Procurações'),
    ('Escritura de Compra e Venda', 'Escrituras'),
    ('Escritura de Doação', 'Escrituras'),
    ('Escritura de Inventário e Partilha', 'Escrituras'),
    ('Escritura de União Estável', 'Escrituras'),
    ('Escritura de Divórcio Consensual', 'Escrituras'),
    ('Escritura Declaratória', 'Escrituras'),
]

FASES = [
    ('Aguardando documentos p/ análise', 'Não', 'Não', 'cinza'),
    ('Pendência de documentos', 'Não', 'Não', 'dourado'),
    ('Em confecção', 'Não', 'Não', 'rosa'),
    ('Em análise pelo Tabelião', 'Não', 'Não', 'azul'),
    ('Em conferência pelo usuário', 'Não', 'Não', 'verde-água'),
    ('Aguardando assinatura', 'Não', 'Não', 'ameixa'),
    ('Finalizado', 'Sim', 'Sim', 'verde'),
    ('Usuário desistiu', 'Sim', 'Não', 'escuro'),
]

# Nomes genericos de proposito: este arquivo vai para um repositorio publico.
# Troque pelos nomes reais direto na aba Listas da planilha (ver secao 12 do guia).
RESPONSAVEIS = ['Escrevente 1', 'Escrevente 2', 'Escrevente 3']
CANAIS = ['Balcão', 'WhatsApp', 'E-mail']
PFPJ = ['PF', 'PJ']

FERIADOS = [
    ('2026-01-01', 'Confraternização Universal'),
    ('2026-02-16', 'Carnaval — segunda (CONFERIR)'),
    ('2026-02-17', 'Carnaval — terça (CONFERIR)'),
    ('2026-04-03', 'Sexta-feira Santa (CONFERIR)'),
    ('2026-04-21', 'Tiradentes'),
    ('2026-05-01', 'Dia do Trabalho'),
    ('2026-06-04', 'Corpus Christi (CONFERIR)'),
    ('2026-09-07', 'Independência do Brasil'),
    ('2026-10-12', 'Nossa Senhora Aparecida'),
    ('2026-11-02', 'Finados'),
    ('2026-11-15', 'Proclamação da República'),
    ('2026-11-20', 'Consciência Negra'),
    ('2026-12-25', 'Natal'),
]

DOCS_BASE = [
    'Documento de identificação com CPF do(s) outorgante(s)',
    'Documento de identificação com CPF do(s) outorgado(s)',
    'Qualificação do(s) outorgante(s)',
    'Qualificação do(s) outorgado(s)',
]

# ---------------------------------------------------------------- estilo
AZUL = '1F3864'
CINZA = 'D9E2F3'
CALC = 'EDEDED'
h_font = Font(bold=True, color='FFFFFF', size=10)
h_fill = PatternFill('solid', fgColor=AZUL)
sub_fill = PatternFill('solid', fgColor=CINZA)
calc_fill = PatternFill('solid', fgColor=CALC)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ================================================================ INÍCIO
ws = wb.active
ws.title = 'Início'
ws.sheet_view.showGridLines = False
ws.column_dimensions['B'].width = 105
linhas_inicio = [
    ('t', 'Controle de Atos — Procurações e Escrituras'),
    ('s', 'Planilha 100% nativa do Google Sheets. Não usa Apps Script.'),
    ('', ''),
    ('h', 'Rotina diária'),
    ('p', '1. Abra a sua Visualização de Filtro: menu Dados › Visualizações de filtro › "Minhas — <seu nome>".'),
    ('p', '2. Ato novo: preencha uma linha na aba PROTOCOLOS. O ID é obrigatório; o Nº DRD entra depois, quando sair.'),
    ('p', '3. Documentos: copie o bloco do tipo de ato na aba LISTAS (colunas R e S), cole no fim da aba CHECKLIST e preencha o ID.'),
    ('p', '4. Ao mexer em um protocolo, atualize "Última atualização" (Ctrl+;) e "Atualizado por".'),
    ('p', '5. Para registrar andamento em texto, use Comentário nativo na linha (Ctrl+Alt+M) — ele grava autor e data/hora sozinho.'),
    ('', ''),
    ('h', 'Regras do setor'),
    ('p', 'Prazo: 5 dias úteis contados da data em que a documentação ficou completa. Sem exceção por tipo de ato.'),
    ('p', 'O prazo só aparece depois de preenchida a coluna "Data doc. completa".'),
    ('p', 'Marcar "Doc. completa?" é decisão de quem confere — o checklist NÃO marca isso sozinho.'),
    ('p', '"Finalizado" exige rascunho digitalizado. A planilha destaca em vermelho, mas não bloqueia.'),
    ('p', 'Nada é excluído: marque "Arquivado?" em vez de apagar a linha.'),
    ('', ''),
    ('h', 'Colunas cinzas são calculadas'),
    ('p', 'Categoria, Prazo, Dias úteis restantes, Situação e % Checklist saem de fórmula. Não digite nelas.'),
    ('', ''),
    ('h', 'Atenção antes do uso real'),
    ('p', 'Confira a aba LISTAS, coluna O: os feriados marcados (CONFERIR) e os feriados estaduais e municipais'),
    ('p', 'ainda precisam ser validados. O prazo de 5 dias úteis depende inteiramente dessa lista.'),
]
r = 1
for kind, txt in linhas_inicio:
    c = ws.cell(row=r, column=2, value=txt)
    if kind == 't':
        c.font = Font(bold=True, size=16, color=AZUL)
        ws.row_dimensions[r].height = 26
    elif kind == 's':
        c.font = Font(italic=True, size=10, color='595959')
    elif kind == 'h':
        c.font = Font(bold=True, size=11, color=AZUL)
        ws.row_dimensions[r].height = 20
    else:
        c.font = Font(size=10)
    c.alignment = left
    r += 1

# ================================================================ PROTOCOLOS
wp = wb.create_sheet('Protocolos')
COLS = [
    ('ID', 12, None),
    ('Nº DRD', 14, None),
    ('Data do protocolo', 14, 'DD/MM/YYYY'),
    ('Cliente', 30, None),
    ('PF/PJ', 8, None),
    ('Telefone', 15, None),
    ('E-mail', 24, None),
    ('Outorgados', 30, None),
    ('Tipo de ato', 28, None),
    ('Categoria', 14, None),            # J calculada
    ('Canal', 11, None),
    ('Responsável', 18, None),
    ('Doc. completa?', 12, None),
    ('Data doc. completa', 15, 'DD/MM/YYYY'),
    ('Prazo', 12, 'DD/MM/YYYY'),        # O calculada
    ('Dias úteis restantes', 12, '0'),  # P calculada
    ('Situação', 13, None),             # Q calculada
    ('Fase', 26, None),
    ('Digitalizado?', 11, None),
    ('Data digitalização', 15, 'DD/MM/YYYY'),
    ('% Checklist', 11, '0%'),          # U calculada
    ('Última atualização', 15, 'DD/MM/YYYY'),
    ('Atualizado por', 16, None),
    ('Observações', 40, None),
    ('Arquivado?', 10, None),
    ('Criado em', 13, 'DD/MM/YYYY'),
]
CALC_COLS = {10, 15, 16, 17, 21}  # J O P Q U
CHECK_COLS = {13, 19, 25}         # M S Y

for i, (nome, larg, fmt) in enumerate(COLS, start=1):
    c = wp.cell(row=1, column=i, value=nome)
    c.font = h_font
    c.fill = h_fill
    c.alignment = center
    c.border = box
    wp.column_dimensions[get_column_letter(i)].width = larg
wp.row_dimensions[1].height = 32
wp.freeze_panes = 'C2'

FERIADOS_REF = 'Listas!$O$2:$O$200'
for r in range(2, LINHAS + 1):
    wp.cell(row=r, column=10,
            value='=IF($I{r}="","",IFERROR(VLOOKUP($I{r},Listas!$A$2:$B$100,2,FALSE),"—"))'.format(r=r))
    wp.cell(row=r, column=15,
            value='=IF($N{r}="","",WORKDAY($N{r},5,{f}))'.format(r=r, f=FERIADOS_REF))
    wp.cell(row=r, column=16,
            value=('=IF($O{r}="","",IF($O{r}>=TODAY(),'
                   'NETWORKDAYS(TODAY(),$O{r},{f})-1,'
                   'NETWORKDAYS(TODAY(),$O{r},{f})+1))').format(r=r, f=FERIADOS_REF))
    wp.cell(row=r, column=17,
            value=('=IF($O{r}="","",IF($R{r}="Finalizado","Concluído",'
                   'IF($R{r}="Usuário desistiu","Encerrado",'
                   'IF($P{r}<0,"Atrasado",IF($P{r}=0,"Vence hoje","No prazo")))))').format(r=r))
    wp.cell(row=r, column=21,
            value=('=IF(COUNTIF(Checklist!$A$2:$A$3000,$A{r})=0,"",'
                   'COUNTIFS(Checklist!$A$2:$A$3000,$A{r},Checklist!$C$2:$C$3000,TRUE)'
                   '/COUNTIF(Checklist!$A$2:$A$3000,$A{r}))').format(r=r))
    for col in CHECK_COLS:
        wp.cell(row=r, column=col, value=False)
    for i, (nome, larg, fmt) in enumerate(COLS, start=1):
        cell = wp.cell(row=r, column=i)
        if fmt:
            cell.number_format = fmt
        if i in CALC_COLS:
            cell.fill = calc_fill
        if i in CHECK_COLS or i in (5, 16, 17, 21):
            cell.alignment = center

val = [
    ('$E$2:$E$%d' % LINHAS, '=Listas!$M$2:$M$5'),
    ('$I$2:$I$%d' % LINHAS, '=Listas!$A$2:$A$100'),
    ('$K$2:$K$%d' % LINHAS, '=Listas!$K$2:$K$20'),
    ('$L$2:$L$%d' % LINHAS, '=Listas!$I$2:$I$50'),
    ('$R$2:$R$%d' % LINHAS, '=Listas!$D$2:$D$100'),
    ('$W$2:$W$%d' % LINHAS, '=Listas!$I$2:$I$50'),
]
for rng, src in val:
    dv = DataValidation(type='list', formula1=src, allow_blank=True, showDropDown=False)
    wp.add_data_validation(dv)
    dv.add(rng)

# ================================================================ CHECKLIST
wc = wb.create_sheet('Checklist')
for i, (nome, larg) in enumerate(
        [('ID Protocolo', 14), ('Documento', 60), ('Entregue?', 11), ('Observação', 40)], start=1):
    c = wc.cell(row=1, column=i, value=nome)
    c.font = h_font
    c.fill = h_fill
    c.alignment = center
    c.border = box
    wc.column_dimensions[get_column_letter(i)].width = larg
wc.row_dimensions[1].height = 32
wc.freeze_panes = 'A2'
for r in range(2, 1002):
    wc.cell(row=r, column=3, value=False).alignment = center

# ================================================================ PAINEL
wl = wb.create_sheet('Painel')
wl.sheet_view.showGridLines = False
for col, larg in (('A', 4), ('B', 36), ('C', 15), ('D', 15), ('E', 15)):
    wl.column_dimensions[col].width = larg

P = 'Protocolos!'


def RNG(c):
    return '{p}${c}$2:${c}${n}'.format(p=P, c=c, n=LINHAS)


def titulo(row, txt):
    wl.cell(row=row, column=2, value=txt).font = Font(bold=True, size=12, color=AZUL)


def par(row, rotulo, formula):
    wl.cell(row=row, column=2, value=rotulo).font = Font(size=10)
    c = wl.cell(row=row, column=3, value=formula)
    c.font = Font(bold=True, size=10)
    c.alignment = center
    c.fill = sub_fill
    c.border = box


titulo(1, 'Painel do setor')
wl.cell(row=2, column=2, value='Próximo ID sugerido').font = Font(size=10)
c = wl.cell(row=2, column=3,
            value='=TEXT(YEAR(TODAY()),"0000")&"-"&TEXT(COUNTA({p}$A$2:$A${n})+1,"0000")'.format(p=P, n=LINHAS))
c.font = Font(bold=True, size=11, color='C00000')
c.alignment = center
c.fill = sub_fill
c.border = box

titulo(4, 'Alertas')
alertas = [
    ('Atrasados', '=COUNTIF({r},"Atrasado")'.format(r=RNG('Q'))),
    ('Vencem hoje', '=COUNTIF({r},"Vence hoje")'.format(r=RNG('Q'))),
    ('Sem responsável', '=SUMPRODUCT(({a}<>"")*({l}=""))'.format(a=RNG('A'), l=RNG('L'))),
    ('Documentação incompleta', '=SUMPRODUCT(({a}<>"")*({m}=FALSE))'.format(a=RNG('A'), m=RNG('M'))),
    ('Finalizado sem digitalização',
     '=SUMPRODUCT(({r}="Finalizado")*({s}=FALSE))'.format(r=RNG('R'), s=RNG('S'))),
    ('Sem Nº DRD', '=SUMPRODUCT(({a}<>"")*({b}=""))'.format(a=RNG('A'), b=RNG('B'))),
]
for i, (rot, f) in enumerate(alertas):
    par(5 + i, rot, f)

titulo(13, 'Por fase (não arquivados)')
for i, (fase, _, _, _) in enumerate(FASES):
    row = 14 + i
    wl.cell(row=row, column=2, value=fase).font = Font(size=10)
    c = wl.cell(row=row, column=3,
                value='=SUMPRODUCT(({r}=$B${row})*({y}=FALSE))'.format(r=RNG('R'), y=RNG('Y'), row=row))
    c.font = Font(size=10)
    c.alignment = center
    c.border = box

titulo(24, 'Por escrevente')
for j, rot in enumerate(['Em andamento', 'Atrasados', 'Vencem hoje']):
    c = wl.cell(row=25, column=3 + j, value=rot)
    c.font = Font(bold=True, size=9)
    c.alignment = center
    c.fill = sub_fill
    c.border = box
for i, nome in enumerate(RESPONSAVEIS):
    row = 26 + i
    wl.cell(row=row, column=2, value=nome).font = Font(size=10)
    ativo = ('=SUMPRODUCT(({a}<>"")*({l}=$B${row})*({r}<>"Finalizado")'
             '*({r}<>"Usuário desistiu")*({y}=FALSE))').format(
        a=RNG('A'), l=RNG('L'), r=RNG('R'), y=RNG('Y'), row=row)
    atras = '=SUMPRODUCT(({l}=$B${row})*({q}="Atrasado"))'.format(l=RNG('L'), q=RNG('Q'), row=row)
    hoje = '=SUMPRODUCT(({l}=$B${row})*({q}="Vence hoje"))'.format(l=RNG('L'), q=RNG('Q'), row=row)
    for j, f in enumerate([ativo, atras, hoje]):
        c = wl.cell(row=row, column=3 + j, value=f)
        c.font = Font(size=10)
        c.alignment = center
        c.border = box

# ================================================================ LISTAS
wi = wb.create_sheet('Listas')
blocos = [
    ('A', ['Tipo de ato', 'Categoria'], [list(t) for t in TIPOS]),
    ('D', ['Fase', 'É final?', 'Exige digitalização?', 'Cor sugerida'], [list(f) for f in FASES]),
    ('I', ['Responsável'], [[n] for n in RESPONSAVEIS]),
    ('K', ['Canal'], [[c] for c in CANAIS]),
    ('M', ['Tipo de pessoa'], [[p] for p in PFPJ]),
    ('O', ['Feriado', 'Descrição'], [[d, t] for d, t in FERIADOS]),
    ('R', ['Tipo (modelo de checklist)', 'Documento'], [['(todos)', d] for d in DOCS_BASE]),
]
for col_ini, cabecalhos, dados in blocos:
    ci = column_index_from_string(col_ini)
    for k, nome in enumerate(cabecalhos):
        c = wi.cell(row=1, column=ci + k, value=nome)
        c.font = h_font
        c.fill = h_fill
        c.alignment = center
        c.border = box
        wi.column_dimensions[get_column_letter(ci + k)].width = 34 if k else 28
    for j, linha in enumerate(dados):
        for k, v in enumerate(linha):
            if col_ini == 'O' and k == 0:
                v = datetime.datetime.strptime(v, '%Y-%m-%d')
            c = wi.cell(row=2 + j, column=ci + k, value=v)
            c.font = Font(size=10)
            if col_ini == 'O' and k == 0:
                c.number_format = 'DD/MM/YYYY'
wi.row_dimensions[1].height = 32
wi.freeze_panes = 'A2'
for col in ('C', 'H', 'J', 'L', 'N', 'Q'):
    wi.column_dimensions[col].width = 3

# ================================================================ FORMATAÇÃO CONDICIONAL
# Gravada no proprio .xlsx para chegar pronta na importacao do Sheets.
# Ordem importa: a primeira regra que casar define o fundo (stopIfTrue).
def f(cor):
    return PatternFill('solid', bgColor=cor)


LINHA = 'A2:Z%d' % LINHAS
regras_linha = [
    # (formula, fill, font)
    ('=$Y2=TRUE', 'EFEFEF', Font(color='9C9C9C', size=10)),
    ('=AND($R2="Finalizado",$S2=FALSE)', 'C00000', Font(color='FFFFFF', bold=True, size=10)),
    ('=$Q2="Atrasado"', 'FFC7CE', Font(color='9C0006', size=10)),
    ('=$Q2="Vence hoje"', 'FFEB9C', Font(color='9C6500', size=10)),
    ('=$Q2="Concluído"', 'C6EFCE', Font(color='006100', size=10)),
]
for formula, cor, fonte in regras_linha:
    wp.conditional_formatting.add(
        LINHA, FormulaRule(formula=[formula], fill=f(cor), font=fonte, stopIfTrue=True))

# Inconsistencia: documentacao marcada como completa com checklist ainda aberto.
wp.conditional_formatting.add(
    LINHA,
    FormulaRule(formula=['=AND($M2=TRUE,$U2<>"",$U2<1)'],
                font=Font(color='BF8F00', bold=True, size=10), stopIfTrue=False))

# Duplicidade de chave.
wp.conditional_formatting.add(
    'A2:A%d' % LINHAS,
    FormulaRule(formula=['=AND($A2<>"",COUNTIF($A$2:$A$%d,$A2)>1)' % LINHAS],
                fill=f('FF9999'), font=Font(color='9C0006', bold=True, size=10)))
wp.conditional_formatting.add(
    'B2:B%d' % LINHAS,
    FormulaRule(formula=['=AND($B2<>"",COUNTIF($B$2:$B$%d,$B2)>1)' % LINHAS],
                fill=f('FF9999'), font=Font(color='9C0006', bold=True, size=10)))

# Cor por escrevente, so na coluna Responsavel.
CORES_ESCREVENTE = ['DDEBF7', 'FCE4D6', 'E4DFEC']
for nome, cor in zip(RESPONSAVEIS, CORES_ESCREVENTE):
    wp.conditional_formatting.add(
        'L2:L%d' % LINHAS,
        FormulaRule(formula=['=$L2="%s"' % nome], fill=f(cor), stopIfTrue=True))

# Checklist: ID que nao existe na aba Protocolos.
wc.conditional_formatting.add(
    'A2:A1001',
    FormulaRule(formula=['=AND($A2<>"",COUNTIF(Protocolos!$A$2:$A$%d,$A2)=0)' % LINHAS],
                fill=f('FF9999'), font=Font(color='9C0006', bold=True, size=10)))
# Checklist: linha ja entregue, esmaecida.
wc.conditional_formatting.add(
    'A2:D1001',
    FormulaRule(formula=['=$C2=TRUE'], font=Font(color='808080', size=10)))

nomes = {
    'Lista_Tipos': 'Listas!$A$2:$A$100',
    'Lista_Fases': 'Listas!$D$2:$D$100',
    'Lista_Responsaveis': 'Listas!$I$2:$I$50',
    'Lista_Canais': 'Listas!$K$2:$K$20',
    'Lista_PFPJ': 'Listas!$M$2:$M$5',
    'Feriados': 'Listas!$O$2:$O$200',
}
for nome, ref in nomes.items():
    wb.defined_names.add(DefinedName(nome, attr_text=ref))

destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Controle-Atos-Cartorio.xlsx')
wb.save(destino)
print('gerado:', os.path.getsize(destino), 'bytes')
